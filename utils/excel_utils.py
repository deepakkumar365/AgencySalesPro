import io
import csv
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill
from app import db
from models import Product, Order, OrderItem

def export_products_to_excel(products):
    """Export products to Excel file using openpyxl"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Products"
    
    # Headers
    headers = ['ID', 'Name', 'Description', 'SKU', 'Price', 'Cost', 'Stock Quantity', 'Category', 'Agency', 'Active', 'Created At']
    ws.append(headers)
    
    # Style headers
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
    
    # Add data
    for product in products:
        row_data = [
            product.id,
            product.name,
            product.description,
            product.sku,
            float(product.price),
            float(product.cost) if product.cost else 0,
            product.stock_quantity,
            product.category,
            product.agency.name,
            product.is_active,
            product.created_at.strftime('%Y-%m-%d %H:%M:%S')
        ]
        ws.append(row_data)
    
    # Auto-size columns
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def import_products_from_excel(file, agency_id, user_role):
    """Import products from Excel or CSV file"""
    try:
        imported = 0
        skipped = 0
        
        if file.filename.lower().endswith('.csv'):
            # Handle CSV file
            content = file.read().decode('utf-8')
            csv_file = io.StringIO(content)
            reader = csv.DictReader(csv_file)
            
            for row in reader:
                # Extract data from row
                name = row.get('Name') or row.get('name') or ''
                sku = row.get('SKU') or row.get('sku') or ''
                price = row.get('Price') or row.get('price') or ''
                
                if not all([name.strip(), sku.strip(), price]):
                    skipped += 1
                    continue
                
                # Check if SKU already exists
                if Product.query.filter_by(sku=sku.strip()).first():
                    skipped += 1
                    continue
                
                # Create product
                try:
                    product = Product(
                        name=name.strip(),
                        description=row.get('Description', '').strip(),
                        sku=sku.strip(),
                        price=float(price),
                        cost=float(row.get('Cost', 0)) if row.get('Cost') else 0,
                        stock_quantity=int(row.get('Stock Quantity', 0)) if row.get('Stock Quantity') else 0,
                        category=row.get('Category', '').strip(),
                        agency_id=agency_id,
                        is_active=True
                    )
                    
                    db.session.add(product)
                    imported += 1
                except (ValueError, TypeError):
                    skipped += 1
                    continue
        
        else:
            # Handle Excel file using openpyxl
            from openpyxl import load_workbook
            
            wb = load_workbook(file)
            ws = wb.active
            
            # Get headers from first row
            headers = []
            for cell in ws[1]:
                headers.append(cell.value)
            
            # Process data rows
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not any(row):  # Skip empty rows
                    continue
                
                # Create row dictionary
                row_dict = {}
                for i, value in enumerate(row):
                    if i < len(headers) and headers[i]:
                        row_dict[headers[i]] = value
                
                # Extract data from row
                name = row_dict.get('Name') or row_dict.get('name') or ''
                sku = row_dict.get('SKU') or row_dict.get('sku') or ''
                price = row_dict.get('Price') or row_dict.get('price') or ''
                
                if not all([str(name).strip(), str(sku).strip(), price]):
                    skipped += 1
                    continue
                
                # Check if SKU already exists
                if Product.query.filter_by(sku=str(sku).strip()).first():
                    skipped += 1
                    continue
                
                # Create product
                try:
                    product = Product(
                        name=str(name).strip(),
                        description=str(row_dict.get('Description', '')).strip(),
                        sku=str(sku).strip(),
                        price=float(price),
                        cost=float(row_dict.get('Cost', 0)) if row_dict.get('Cost') else 0,
                        stock_quantity=int(row_dict.get('Stock Quantity', 0)) if row_dict.get('Stock Quantity') else 0,
                        category=str(row_dict.get('Category', '')).strip(),
                        agency_id=agency_id,
                        is_active=True
                    )
                    
                    db.session.add(product)
                    imported += 1
                except (ValueError, TypeError):
                    skipped += 1
                    continue
        
        db.session.commit()
        
        return {
            'success': True,
            'imported': imported,
            'skipped': skipped
        }
        
    except Exception as e:
        db.session.rollback()
        return {
            'success': False,
            'message': str(e)
        }

def export_orders_to_excel(orders):
    """Export orders to Excel file using openpyxl"""
    wb = Workbook()
    
    # Orders details sheet
    ws_details = wb.active
    ws_details.title = "Order Details"
    
    # Headers for details
    detail_headers = [
        'Order ID', 'Order Number', 'Customer', 'Customer Email', 'Customer Phone',
        'Location', 'Agency', 'Salesperson', 'Product Name', 'Product SKU',
        'Quantity', 'Unit Price', 'Total Price', 'Order Status', 'Order Total',
        'Discount', 'Tax', 'Order Date', 'Delivery Date', 'Notes'
    ]
    ws_details.append(detail_headers)
    
    # Style headers
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    for cell in ws_details[1]:
        cell.font = header_font
        cell.fill = header_fill
    
    # Add order details data
    for order in orders:
        for item in order.order_items:
            row_data = [
                order.id,
                order.order_number,
                order.customer.name,
                order.customer.email,
                order.customer.phone,
                order.customer.location.name,
                order.agency.name,
                order.salesperson.full_name,
                item.product.name,
                item.product.sku,
                item.quantity,
                float(item.unit_price),
                float(item.total_price),
                order.status,
                float(order.total_amount),
                float(order.discount),
                float(order.tax),
                order.order_date.strftime('%Y-%m-%d %H:%M:%S'),
                order.delivery_date.strftime('%Y-%m-%d') if order.delivery_date else '',
                order.notes
            ]
            ws_details.append(row_data)
    
    # Auto-size columns for details
    for column in ws_details.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_details.column_dimensions[column_letter].width = adjusted_width
    
    # Order summary sheet
    ws_summary = wb.create_sheet("Order Summary")
    summary_headers = [
        'Order Number', 'Customer', 'Agency', 'Salesperson', 'Status',
        'Total Amount', 'Order Date', 'Items Count'
    ]
    ws_summary.append(summary_headers)
    
    # Style summary headers
    for cell in ws_summary[1]:
        cell.font = header_font
        cell.fill = header_fill
    
    # Add summary data
    order_summary = {}
    for order in orders:
        if order.id not in order_summary:
            row_data = [
                order.order_number,
                order.customer.name,
                order.agency.name,
                order.salesperson.full_name,
                order.status,
                float(order.total_amount),
                order.order_date.strftime('%Y-%m-%d %H:%M:%S'),
                len(order.order_items)
            ]
            ws_summary.append(row_data)
            order_summary[order.id] = True
    
    # Auto-size columns for summary
    for column in ws_summary.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_summary.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output